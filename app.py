
import streamlit as st
import pandas as pd
import sqlite3
from datetime import date, datetime
from pathlib import Path

APP_DIR = Path(__file__).parent
DB_PATH = APP_DIR / "worklog.db"

COLUMNS = [
    ("order_type", "Order Type"),
    ("work_date", "Date"),
    ("job_number", "Job Number"),
    ("job_type", "Job Type"),
    ("vehicle_description", "Vehicle Description"),
    ("vehicle_reg", "Vehicle Reg"),
    ("from_loc", "From"),
    ("to_loc", "To"),
    ("job_amount", "Job Amount"),
    ("expenses", "Expenses"),
    ("expense_amount", "Expense Amount"),
    ("comments", "Comments"),
    ("auth_code", "AUTH CODE"),
    ("job_status", "Job Status"),
]

def conn():
    return sqlite3.connect(DB_PATH, check_same_thread=False)

def init_db():
    with conn() as con:
        con.execute("""
        CREATE TABLE IF NOT EXISTS entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_type TEXT,
            work_date TEXT,
            job_number TEXT,
            job_type TEXT,
            vehicle_description TEXT,
            vehicle_reg TEXT,
            from_loc TEXT,
            to_loc TEXT,
            job_amount REAL,
            expenses TEXT,
            expense_amount TEXT,
            comments TEXT,
            auth_code TEXT,
            job_status TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now'))
        )
        """)
        con.execute("CREATE INDEX IF NOT EXISTS idx_entries_date ON entries(work_date)")
        con.execute("CREATE INDEX IF NOT EXISTS idx_entries_jobnum ON entries(job_number)")
        con.commit()

def parse_date_any(x):
    if x is None or (isinstance(x, float) and pd.isna(x)) or (isinstance(x, str) and not x.strip()):
        return None
    if isinstance(x, (datetime, date)):
        return x.date().isoformat()
    try:
        # common ISO / UK-ish inputs
        return pd.to_datetime(x, dayfirst=True, errors="coerce").date().isoformat()
    except Exception:
        return None

def to_float(x):
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if not s:
        return None
    try:
        return float(s)
    except Exception:
        return None

def load_entries():
    with conn() as con:
        df = pd.read_sql_query("SELECT * FROM entries ORDER BY work_date DESC, id DESC", con)
    if not df.empty:
        df["work_date"] = pd.to_datetime(df["work_date"], errors="coerce").dt.date
    return df

def insert_entry(values: dict):
    cols = [c[0] for c in COLUMNS]
    placeholders = ",".join(["?"] * len(cols))
    sql = f"INSERT INTO entries ({','.join(cols)}) VALUES ({placeholders})"
    with conn() as con:
        con.execute(sql, [values.get(k) for k in cols])
        con.commit()

def update_entries(df_changed: pd.DataFrame):
    # df_changed must include id
    allowed = set(["id"] + [c[0] for c in COLUMNS])
    df_changed = df_changed[[c for c in df_changed.columns if c in allowed]].copy()
    with conn() as con:
        for _, row in df_changed.iterrows():
            rid = int(row["id"])
            sets, params = [], []
            for col, _label in COLUMNS:
                if col in row.index:
                    sets.append(f"{col}=?")
                    v = row[col]
                    if col == "work_date":
                        v = parse_date_any(v)
                    elif col == "job_amount":
                        v = to_float(v)
                    params.append(v if pd.notna(v) else None)
            params.append(rid)
            con.execute(f"UPDATE entries SET {', '.join(sets)}, updated_at=datetime('now') WHERE id=?", params)
        con.commit()

def delete_ids(ids):
    if not ids:
        return
    with conn() as con:
        con.executemany("DELETE FROM entries WHERE id=?", [(int(i),) for i in ids])
        con.commit()

def import_excel(file):
    # Reads every sheet, expects headers in first row similar to your template.
    import openpyxl
    wb = openpyxl.load_workbook(file, data_only=True)
    all_rows = []
    for name in wb.sheetnames:
        ws = wb[name]
        headers = []
        for c in range(1, ws.max_column + 1):
            headers.append(ws.cell(1, c).value)
        if not headers or headers[0] is None:
            continue
        for r in range(2, ws.max_row + 1):
            row = {headers[c-1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
            # skip fully empty rows
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in row.values()):
                continue
            all_rows.append(row)

    if not all_rows:
        return 0

    df = pd.DataFrame(all_rows)
    # Map your Excel headers to our DB cols
    mapping = {
        "order type": "order_type",
        "Date": "work_date",
        "Job Number": "job_number",
        "Job type": "job_type",
        "Vehicle description": "vehicle_description",
        "Vehicle Reg": "vehicle_reg",
        "From": "from_loc",
        "To": "to_loc",
        "Job Amount": "job_amount",
        "Expenses": "expenses",
        "Amount": "expense_amount",
        "Comments": "comments",
        "AUTH CODE": "auth_code",
        "Job Status": "job_status",
    }

    # Normalize columns (case-insensitive where relevant)
    cols_norm = {c: (str(c).strip() if c is not None else "") for c in df.columns}
    df = df.rename(columns={c: mapping.get(cols_norm[c], cols_norm[c]) for c in df.columns})

    needed = [c[0] for c in COLUMNS]
    for col in needed:
        if col not in df.columns:
            df[col] = None

    df = df[needed].copy()
    df["work_date"] = df["work_date"].apply(parse_date_any)
    df["job_amount"] = df["job_amount"].apply(to_float)

    with conn() as con:
        con.executemany(
            f"INSERT INTO entries ({','.join(needed)}) VALUES ({','.join(['?']*len(needed))})",
            df.where(pd.notna(df), None).values.tolist()
        )
        con.commit()
    return len(df)

st.set_page_config(page_title="Work Log", layout="wide")
st.title("Work Log (simple frontend)")

init_db()

with st.sidebar:
    st.header("Filters")
    df_all = load_entries()

    if df_all.empty:
        st.info("No data yet. Import your Excel file or add entries.")
        min_d = date.today()
        max_d = date.today()
    else:
        min_d = df_all["work_date"].min() if df_all["work_date"].notna().any() else date.today()
        max_d = df_all["work_date"].max() if df_all["work_date"].notna().any() else date.today()

    date_range = st.date_input("Date range", value=(min_d, max_d))
    status = st.multiselect("Job Status", sorted([x for x in df_all["job_status"].dropna().unique()]) if not df_all.empty else [])
    job_type = st.multiselect("Job Type", sorted([x for x in df_all["job_type"].dropna().unique()]) if not df_all.empty else [])
    search = st.text_input("Search (job #, reg, places, comments)")

    st.divider()
    st.subheader("Import")
    up = st.file_uploader("Import Excel (.xlsx)", type=["xlsx"])
    if up is not None:
        count = import_excel(up)
        st.success(f"Imported {count} rows.")
        st.rerun()

# Apply filters
df = df_all.copy()
if not df.empty:
    if isinstance(date_range, tuple) and len(date_range) == 2:
        start, end = date_range
        df = df[(df["work_date"] >= start) & (df["work_date"] <= end)]
    if status:
        df = df[df["job_status"].isin(status)]
    if job_type:
        df = df[df["job_type"].isin(job_type)]
    if search.strip():
        s = search.strip().lower()
        hay = (
            df["job_number"].fillna("").astype(str) + " " +
            df["vehicle_reg"].fillna("").astype(str) + " " +
            df["from_loc"].fillna("").astype(str) + " " +
            df["to_loc"].fillna("").astype(str) + " " +
            df["comments"].fillna("").astype(str)
        ).str.lower()
        df = df[hay.str.contains(s, na=False)]

st.subheader("Entries")

if df.empty:
    st.write("No rows match your filters yet.")
else:
    # Editable grid
    display_cols = ["id"] + [c[0] for c in COLUMNS]
    df_show = df[display_cols].copy()
    # Make date readable
    df_show["work_date"] = pd.to_datetime(df_show["work_date"], errors="coerce").dt.date

    edited = st.data_editor(
        df_show,
        use_container_width=True,
        num_rows="fixed",
        key="editor",
        column_config={
            "id": st.column_config.NumberColumn("ID", disabled=True),
            "work_date": st.column_config.DateColumn("Date"),
            "job_amount": st.column_config.NumberColumn("Job Amount"),
        },
        hide_index=True
    )

    colA, colB, colC = st.columns([1,1,2])
    with colA:
        if st.button("Save changes", type="primary"):
            # find rows that changed (cheap approach: update all visible rows)
            update_entries(edited)
            st.success("Saved.")
            st.rerun()
    with colB:
        del_ids = st.multiselect("Delete IDs", options=edited["id"].tolist())
        if st.button("Delete selected", type="secondary"):
            delete_ids(del_ids)
            st.warning(f"Deleted {len(del_ids)} row(s).")
            st.rerun()

st.divider()
st.subheader("Add a new entry")

with st.form("add_form", clear_on_submit=True):
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        order_type = st.selectbox("Order Type", options=["Day", "CarryOver", "Other"], index=0)
        work_date = st.date_input("Date", value=date.today())
        job_number = st.text_input("Job Number")
    with c2:
        job_type_in = st.text_input("Job Type")
        job_status_in = st.text_input("Job Status")
        auth_code = st.text_input("AUTH CODE")
    with c3:
        vehicle_description = st.text_input("Vehicle Description")
        vehicle_reg = st.text_input("Vehicle Reg")
        job_amount = st.number_input("Job Amount", value=0.0, step=0.5)
    with c4:
        from_loc = st.text_input("From")
        to_loc = st.text_input("To")
        expenses = st.text_area("Expenses", height=68)
        expense_amount = st.text_input("Expense Amount")

    comments = st.text_area("Comments")

    submitted = st.form_submit_button("Add entry")
    if submitted:
        values = dict(
            order_type=order_type,
            work_date=work_date.isoformat(),
            job_number=job_number.strip() or None,
            job_type=job_type_in.strip() or None,
            vehicle_description=vehicle_description.strip() or None,
            vehicle_reg=vehicle_reg.strip() or None,
            from_loc=from_loc.strip() or None,
            to_loc=to_loc.strip() or None,
            job_amount=float(job_amount) if job_amount is not None else None,
            expenses=expenses.strip() or None,
            expense_amount=expense_amount.strip() or None,
            comments=comments.strip() or None,
            auth_code=auth_code.strip() or None,
            job_status=job_status_in.strip() or None,
        )
        insert_entry(values)
        st.success("Added.")
        st.rerun()
